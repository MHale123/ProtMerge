#!/usr/bin/env python3
"""
ProtMerge - Protein Analysis Tool
Version: 1.2.0
Created by Matthew Hale
"""

import sys
import logging
import time
from pathlib import Path
from gui_main import ProtMergeGUI
from data_handler import DataHandler
from analyzers import AnalyzerManager
from excel_formatter import ExcelFormatter

class ProtMerge:
    """Main ProtMerge application class"""
    
    def __init__(self):
        self.data_handler = DataHandler()
        self.analyzer_manager = AnalyzerManager()
        self.excel_formatter = ExcelFormatter()
        self.logger = logging.getLogger(__name__)
        self.analysis_summary = None
        self.logger.info("ProtMerge v1.2.0 initialized")
        
    def run_gui(self):
        """Launch the GUI interface"""
        gui = ProtMergeGUI(self)
        gui.run()
    
    def run_analysis(self, input_file, sheet_name, column_index, options, progress_callback=None):
        """Run complete analysis pipeline with flexible dependency handling"""
        try:
            self.logger.info(f"Starting analysis for {Path(input_file).name}")
    
            # Load data
            if progress_callback:
                progress_callback(0, "Loading data", f"Reading {Path(input_file).name}")
    
            data = self.data_handler.load_excel_data(
                input_file, sheet_name, column_index, options.get('safe_mode', True)
            )
    
            protein_count = len(data['results'])
            self.logger.info(f"Loaded {protein_count} proteins for analysis")
            
            # Determine input type and required analyses
            using_gene_ids = options.get('use_gene_ids', False)
            needs_uniprot_conversion = self._needs_uniprot_conversion(options, using_gene_ids)
            
            self.logger.info(f"Input type: {'Gene IDs' if using_gene_ids else 'UniProt IDs'}")
            self.logger.info(f"UniProt conversion needed: {needs_uniprot_conversion}")
            
            # Phase 1: Human-specific analysis (for gene IDs only)
            if using_gene_ids and self._has_human_analyses(options):
                if progress_callback:
                    progress_callback(5, "Running human protein analysis", "Analyzing human-specific databases")
                
                self.logger.info("Running human-specific analysis before any conversion...")
                data = self.analyzer_manager.run_human_protein_analysis(data, options, progress_callback)
                
                # Log human analysis results
                self._log_human_analysis_results(data['results'], options, protein_count)
            
            # Phase 2: Gene to UniProt conversion (only if needed for downstream analyses)
            if using_gene_ids and needs_uniprot_conversion:
                start_progress = 25 if self._has_human_analyses(options) else 5
                if progress_callback:
                    progress_callback(start_progress, "Converting Gene IDs", "Converting gene names to UniProt IDs for downstream analyses")
        
                self.logger.info("Converting gene IDs to UniProt IDs for UniProt-dependent analyses...")
                data = self.analyzer_manager.run_gene_conversion(data, progress_callback)
        
                # Count successful conversions
                converted_count = sum(1 for _, row in data['results'].iterrows() 
                                    if row.get('Original_Gene_ID', '') and 
                                        row.get('UniProt_ID', '') != row.get('Original_Gene_ID', ''))
        
                self.logger.info(f"Gene conversion: {converted_count}/{protein_count} successful")
            
            # Phase 3: UniProt-dependent analyses (only if requested and conversion successful/not needed)
            uniprot_analyses_requested = self._has_uniprot_dependent_analyses(options)
            
            if uniprot_analyses_requested:
                # Check if we have UniProt IDs to work with
                has_uniprot_ids = self._has_valid_uniprot_ids(data['results'], using_gene_ids)
                
                if has_uniprot_ids:
                    # Calculate starting progress
                    if using_gene_ids and needs_uniprot_conversion:
                        start_progress = 45 if self._has_human_analyses(options) else 25
                    elif using_gene_ids:
                        start_progress = 25 if self._has_human_analyses(options) else 5
                    else:
                        start_progress = 5
                
                    if progress_callback:
                        progress_callback(start_progress, "Running UniProt-dependent analyses", "Starting protein data collection")
            
                    results = self.analyzer_manager.run_uniprot_analyses(data, options, progress_callback)
                else:
                    self.logger.warning("No valid UniProt IDs available for UniProt-dependent analyses")
                    results = data['results']
                    # Still initialize columns for consistency
                    self.analyzer_manager._initialize_uniprot_columns(results, options)
            else:
                self.logger.info("No UniProt-dependent analyses requested")
                results = data['results']
    
            # Calculate analysis summary
            self.analysis_summary = self._calculate_analysis_summary(results, options)
    
            # Save results
            if progress_callback:
                progress_callback(98, "Saving results", "Creating Excel file")
    
            output_file = self.excel_formatter.save_results(input_file, results, options)
    
            # Log completion summary
            self._log_completion_summary(output_file, results, options)
    
            if progress_callback:
                progress_callback(100, "Analysis complete", "Results saved successfully")
    
            return output_file
    
        except Exception as e:
            self.logger.error(f"Analysis pipeline failed: {e}")
            raise
    
    def _needs_uniprot_conversion(self, options, using_gene_ids):
        """Determine if UniProt conversion is needed based on requested analyses"""
        if not using_gene_ids:
            return False  # Already have UniProt IDs
        
        # Check if any UniProt-dependent analyses are requested
        return self._has_uniprot_dependent_analyses(options)
    
    def _has_uniprot_dependent_analyses(self, options):
        """Check if any UniProt-dependent analyses are requested"""
        uniprot_dependent = [
            'uniprot',      # UniProt data itself
            'protparam',    # Needs sequence from UniProt
            'blast',        # Needs sequence from UniProt
            'pdb_search'    # Needs UniProt ID for PDB search
        ]
        
        return any(options.get(analysis, False) for analysis in uniprot_dependent)
    
    def _has_human_analyses(self, options):
        """Check if human-specific analyses are requested"""
        return options.get('compartments', False) or options.get('hpa', False)
    
    def _has_valid_uniprot_ids(self, results, using_gene_ids):
        """Check if we have valid UniProt IDs to work with"""
        if using_gene_ids:
            # Check if conversion was successful for at least some entries
            valid_count = sum(1 for _, row in results.iterrows() 
                            if row.get('UniProt_ID', '') and 
                               row.get('UniProt_ID', '') != row.get('Original_Gene_ID', '') and
                               row.get('UniProt_ID', '') not in ['', 'NO VALUE FOUND'])
            return valid_count > 0
        else:
            # Already using UniProt IDs
            valid_count = sum(1 for _, row in results.iterrows() 
                            if row.get('UniProt_ID', '') and 
                               row.get('UniProt_ID', '') not in ['', 'NO VALUE FOUND'])
            return valid_count > 0
    
    def _log_human_analysis_results(self, results, options, protein_count):
        """Log human analysis results - Fixed to use correct column names"""
        if options.get('compartments', False):
            # Use the correct column name from human_protein_analyzer.py
            compartments_complete = sum(1 for _, row in results.iterrows() 
                                    if self._is_data_complete(row.get('compartments_primary_location', '')))
            self.logger.info(f"COMPARTMENTS analysis: {compartments_complete}/{protein_count} successful")
    
        if options.get('hpa', False):
            # Use the correct column name from human_protein_analyzer.py
            hpa_complete = sum(1 for _, row in results.iterrows() 
                            if self._is_data_complete(row.get('hpa_primary_tissue', '')))
            self.logger.info(f"HPA analysis: {hpa_complete}/{protein_count} successful")
    
    def get_analysis_summary(self):
        """Get analysis summary for completion dialog"""
        return self.analysis_summary or {}
    
    def _calculate_analysis_summary(self, results, options):
        """Calculate analysis summary statistics - Fixed to use correct column names"""
        try:
            total_proteins = len(results)
            summary = {'total_proteins': total_proteins}
    
            # Count data completeness only for analyses that were actually requested
        
            # UniProt analysis - check for function data (core UniProt field)
            if options.get('uniprot', False):
                uniprot_complete = sum(1 for _, row in results.iterrows() 
                                    if self._is_data_complete(row.get('function', '')))
                summary['uniprot_complete'] = uniprot_complete
                summary['uniprot_percent'] = (uniprot_complete / total_proteins) * 100 if total_proteins > 0 else 0
        
            # ProtParam analysis - check for molecular weight data
            if options.get('protparam', False):
                protparam_complete = sum(1 for _, row in results.iterrows() 
                                    if self._is_data_complete(row.get('mw', '')))
                summary['protparam_complete'] = protparam_complete
                summary['protparam_percent'] = (protparam_complete / total_proteins) * 100 if total_proteins > 0 else 0
            
            # BLAST analysis - check for identity data
            if options.get('blast', False):
                blast_complete = sum(1 for _, row in results.iterrows() 
                                if self._is_data_complete(row.get('identity', '')))
                summary['blast_complete'] = blast_complete
                summary['blast_percent'] = (blast_complete / total_proteins) * 100 if total_proteins > 0 else 0
            
            # PDB analysis - check for structure count (should be > 0)
            if options.get('pdb_search', False):
                pdb_complete = sum(1 for _, row in results.iterrows() 
                                if self._is_data_complete(row.get('structure_count', ''), check_zero=True))
                summary['pdb_complete'] = pdb_complete
                summary['pdb_percent'] = (pdb_complete / total_proteins) * 100 if total_proteins > 0 else 0
    
            # FIXED: Human protein analyses - use correct column names
            if options.get('compartments', False):
                # Check the actual COMPARTMENTS column name from human_protein_analyzer.py
                compartments_complete = sum(1 for _, row in results.iterrows() 
                                        if self._is_data_complete(row.get('compartments_primary_location', '')))
                summary['compartments_complete'] = compartments_complete
                summary['compartments_percent'] = (compartments_complete / total_proteins) * 100 if total_proteins > 0 else 0
    
            if options.get('hpa', False):
                # Check the actual HPA column name from human_protein_analyzer.py
                hpa_complete = sum(1 for _, row in results.iterrows() 
                                if self._is_data_complete(row.get('hpa_primary_tissue', '')))
                summary['hpa_complete'] = hpa_complete
                summary['hpa_percent'] = (hpa_complete / total_proteins) * 100 if total_proteins > 0 else 0
    
            # Set to 0 for analyses that weren't requested (to avoid showing in completion dialog)
            if not options.get('uniprot', False):
                summary['uniprot_complete'] = 0
                summary['uniprot_percent'] = 0
            if not options.get('protparam', False):
                summary['protparam_complete'] = 0
                summary['protparam_percent'] = 0
            if not options.get('blast', False):
                summary['blast_complete'] = 0
                summary['blast_percent'] = 0
            if not options.get('pdb_search', False):
                summary['pdb_complete'] = 0
                summary['pdb_percent'] = 0
            if not options.get('compartments', False):
                summary['compartments_complete'] = 0
                summary['compartments_percent'] = 0
            if not options.get('hpa', False):
                summary['hpa_complete'] = 0
                summary['hpa_percent'] = 0
    
            return summary
    
        except Exception as e:
            self.logger.error(f"Error calculating analysis summary: {e}")
            return {
                'total_proteins': len(results) if results is not None else 0,
                'uniprot_complete': 0,
                'uniprot_percent': 0,
                'protparam_complete': 0,
                'protparam_percent': 0,
                'blast_complete': 0,
                'blast_percent': 0,
                'pdb_complete': 0,
                'pdb_percent': 0,
                'compartments_complete': 0,
                'compartments_percent': 0,
                'hpa_complete': 0,
                'hpa_percent': 0
            }
    
    def _is_data_complete(self, value, check_zero=False):
        """Check if data value indicates completion"""
        if value is None:
            return False
        
        value_str = str(value).strip().upper()
        
        # Check for no value indicators
        if value_str in ['', 'NO VALUE FOUND', 'NAN', 'NONE', 'N/A']:
            return False
        
        # For PDB structure count, check if it's a positive number
        if check_zero:
            try:
                num_value = float(value_str)
                return num_value > 0
            except (ValueError, TypeError):
                return False
        
        return True
    
    def _log_completion_summary(self, output_file, results, options):
        """Log analysis completion summary - Fixed to only show analyses that were actually run"""
        try:
            summary = self.analysis_summary
        
            self.logger.info("=" * 60)
            self.logger.info("ANALYSIS COMPLETION SUMMARY")
            self.logger.info("=" * 60)
            self.logger.info(f"Total proteins analyzed: {summary['total_proteins']}")
            
            # Create lists to track what was actually run vs requested
            analyses_run = []
            analyses_skipped = []
        
            # Check each analysis type
            analysis_checks = [
                ('compartments', 'COMPARTMENTS', 'compartments_complete', 'compartments_percent'),
                ('hpa', 'Human Protein Atlas', 'hpa_complete', 'hpa_percent'),
                ('uniprot', 'UniProt', 'uniprot_complete', 'uniprot_percent'),
                ('protparam', 'ProtParam', 'protparam_complete', 'protparam_percent'),
                ('blast', 'BLAST', 'blast_complete', 'blast_percent'),
                ('pdb_search', 'PDB Structures', 'pdb_complete', 'pdb_percent')
            ]
            
            for option_key, display_name, complete_key, percent_key in analysis_checks:
                if options.get(option_key, False):
                    complete_count = summary[complete_key]
                    percent = summary[percent_key]
                    self.logger.info(f"{display_name}: {complete_count}/{summary['total_proteins']} ({percent:.1f}%)")
                    analyses_run.append(display_name)
                    
                    if complete_count == 0:
                        analyses_skipped.append(f"{display_name} (0 successful)")
                else:
                    # Analysis was not requested, don't log it
                    pass
            
            # Summary of what was actually executed
            if analyses_run:
                self.logger.info(f"Analyses executed: {', '.join(analyses_run)}")
            else:
                self.logger.info("No analyses were executed")
            
            # Warn about analyses that produced no data
            if analyses_skipped:
                self.logger.warning(f"Analyses with no successful results: {', '.join(analyses_skipped)}")
        
            if output_file:
                self.logger.info(f"Results saved to: {output_file}")
            
                # List actually created sheets
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_file)
                    actual_sheets = [ws for ws in wb.sheetnames if any(name in ws for name in ['ProtMerge', 'Amino', 'PDB', 'Human', 'Similarity'])]
                    self.logger.info(f"Excel sheets created: {', '.join(actual_sheets)}")
                except Exception:
                    self.logger.info("Excel file created (sheet details unavailable)")
        
            self.logger.info("=" * 60)
        
        except Exception as e:
            self.logger.error(f"Error generating completion summary: {e}")
    
    def check_dependencies(self):
        """Check if all dependencies are available"""
        missing_deps = []
        
        # Check core dependencies
        core_deps = ['pandas', 'requests', 'openpyxl', 'lxml']
        for dep in core_deps:
            try:
                __import__(dep)
            except ImportError:
                missing_deps.append(dep)
        
        if missing_deps:
            self.logger.error(f"Missing core dependencies: {', '.join(missing_deps)}")
            return False
        
        self.logger.info("All core dependencies available")
        return True

def setup_logging():
    """Setup logging configuration"""
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    
    # Create logs directory
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            logging.FileHandler(log_dir / f"protmerge_{int(time.time())}.log"),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    # Reduce noise from external libraries
    logging.getLogger('urllib3').setLevel(logging.WARNING)
    logging.getLogger('requests').setLevel(logging.WARNING)

def main():
    """Main entry point"""
    try:        
        # Setup logging
        setup_logging()
        logger = logging.getLogger(__name__)
        
        logger.info("Starting ProtMerge v1.2.0 - Protein Analysis Tool")
        
        # Initialize application
        app = ProtMerge()
        
        # Check dependencies
        if not app.check_dependencies():
            logger.error("Missing required dependencies. Please install them and try again.")
            sys.exit(1)
        
        # Launch GUI
        logger.info("Launching graphical interface...")
        app.run_gui()
        logger.info("ProtMerge session ended")
        
    except KeyboardInterrupt:
        print("\nProtMerge interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"ProtMerge failed to start: {e}")
        logging.error(f"Fatal error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()