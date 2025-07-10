"""
similarity_gui.py - Enhanced Version with Sliders for ProtMerge v1.2.0
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
import threading
import time
from pathlib import Path
import logging

# Import our rewritten similarity analyzer
from similarity_analyzer import SimilarityAnalyzer, SimilarityPresets, run_similarity_analysis

# Import Theme from main GUI (with fallback)
try:
    from gui_main import Theme, ModernButton
except ImportError:
    # Fallback theme if gui_main not available
    class Theme:
        BG = "#1e1e1e"
        SECONDARY = "#2d2d2d"
        TERTIARY = "#3c3c3c"
        TEXT = "#ffffff"
        TEXT_MUTED = "#b0b0b0"
        CYAN = "#00bcd4"
        CYAN_HOVER = "#00acc1"
        GREEN = "#4caf50"
        RED = "#f44336"
        PURPLE = "#9c27b0"
    
    class ModernButton(tk.Button):
        def __init__(self, parent, text, command=None, style="primary", **kwargs):
            colors = {
                "primary": ("#00bcd4", "#00acc1", "white"),
                "success": ("#4caf50", "#45a049", "white"),
                "danger": ("#f44336", "#da190b", "white"),
                "secondary": ("#3c3c3c", "#4a4a4a", "#ffffff")
            }
            
            bg_normal, bg_hover, fg_color = colors.get(style, colors["primary"])
            
            super().__init__(
                parent, text=text, command=command,
                bg=bg_normal, fg=fg_color,
                font=("Segoe UI", 11, "bold"),
                relief="flat", borderwidth=0,
                padx=20, pady=10, cursor="hand2",
                activebackground=bg_hover,
                activeforeground=fg_color, **kwargs
            )
            
            self.bg_normal = bg_normal
            self.bg_hover = bg_hover
            
            self.bind("<Enter>", lambda e: self.config(bg=bg_hover))
            self.bind("<Leave>", lambda e: self.config(bg=bg_normal))


class SimilarityProgressDialog:
    """Progress dialog for similarity analysis"""
    
    def __init__(self, parent):
        self.parent = parent
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Similarity Analysis")
        self.dialog.geometry("400x180")
        self.dialog.configure(bg=Theme.BG)
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center on parent
        self.dialog.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - 200
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - 90
        self.dialog.geometry(f"+{x}+{y}")
        
        self._create_content()
        self.cancelled = False
    
    def _create_content(self):
        """Create dialog content"""
        # Header
        header_frame = tk.Frame(self.dialog, bg=Theme.CYAN, height=50)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🔬 Running Similarity Analysis",
                font=("Segoe UI", 14, "bold"),
                fg="white", bg=Theme.CYAN).pack(expand=True, pady=10)
        
        # Content area
        content_frame = tk.Frame(self.dialog, bg=Theme.BG)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        self.status_label = tk.Label(content_frame, text="Initializing...",
                                   font=("Segoe UI", 10),
                                   fg=Theme.TEXT_MUTED, bg=Theme.BG)
        self.status_label.pack(pady=(0, 10))
        
        self.progress = ttk.Progressbar(content_frame, mode='determinate', length=300)
        self.progress.pack(pady=(0, 15))
        
        button_frame = tk.Frame(content_frame, bg=Theme.BG)
        button_frame.pack(fill=tk.X)
        
        self.cancel_btn = ModernButton(button_frame, "Cancel", self._cancel, "secondary")
        self.cancel_btn.pack(anchor=tk.CENTER)
    
    def update_progress(self, percentage, text):
        try:
            self.progress['value'] = max(0, min(100, percentage))
            self.status_label.config(text=text)
            self.dialog.update()
        except tk.TclError:
            pass
    
    def _cancel(self):
        self.cancelled = True
        self.close()
    
    def close(self):
        try:
            self.dialog.grab_release()
            self.dialog.destroy()
        except tk.TclError:
            pass


class SimilarityResultsViewer:
    """Results viewer for similarity analysis"""
    
    def __init__(self, parent, results_df, central_protein):
        self.parent = parent
        self.results_df = results_df
        self.central_protein = central_protein
        
        self.window = tk.Toplevel(parent)
        self.window.title(f"Similarity Results - Central: {central_protein}")
        self.window.geometry("900x700")
        self.window.configure(bg=Theme.BG)
        self.window.transient(parent)
        
        x = parent.winfo_rootx() + 50
        y = parent.winfo_rooty() + 50
        self.window.geometry(f"+{x}+{y}")
        
        self._create_interface()
    
    def _create_interface(self):
        # Header
        header_frame = tk.Frame(self.window, bg=Theme.CYAN, height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        header_content = tk.Frame(header_frame, bg=Theme.CYAN)
        header_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(header_content, text="🔬 Similarity Results",
                font=("Segoe UI", 16, "bold"),
                fg="white", bg=Theme.CYAN).pack(side=tk.LEFT)
        
        tk.Label(header_content, text=f"Central: {self.central_protein}",
                font=("Segoe UI", 12),
                fg="white", bg=Theme.CYAN).pack(side=tk.RIGHT)
        
        # Results table
        self._create_results_table()
        
        # Summary and buttons
        self._create_summary()
        self._create_buttons()
    
    def _create_results_table(self):
        table_frame = tk.Frame(self.window, bg=Theme.BG)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tree_frame = tk.Frame(table_frame, bg=Theme.BG)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('Rank', 'Protein ID', 'Overall Similarity', 'Data Quality')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.tree.heading(col, text=col)
        
        self.tree.column('Rank', width=80, anchor=tk.CENTER)
        self.tree.column('Protein ID', width=150, anchor=tk.CENTER)
        self.tree.column('Overall Similarity', width=120, anchor=tk.CENTER)
        self.tree.column('Data Quality', width=100, anchor=tk.CENTER)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self._populate_table()
    
    def _populate_table(self):
        if self.results_df.empty:
            self.tree.insert('', 'end', values=('', 'No results', 'No data available', ''))
            return
        
        for idx, row in self.results_df.head(50).iterrows():
            rank = idx + 1
            protein_id = row['protein_id']
            similarity = f"{row['overall_similarity']:.3f}"
            quality = f"{row['data_quality']:.2f}"
            
            self.tree.insert('', 'end', values=(rank, protein_id, similarity, quality))
    
    def _create_summary(self):
        summary_frame = tk.Frame(self.window, bg=Theme.SECONDARY)
        summary_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        title_frame = tk.Frame(summary_frame, bg=Theme.TERTIARY, height=30)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="📊 Summary Statistics",
                font=("Segoe UI", 12, "bold"),
                fg=Theme.TEXT, bg=Theme.TERTIARY).pack(side=tk.LEFT, padx=15, pady=5)
        
        stats_frame = tk.Frame(summary_frame, bg=Theme.SECONDARY)
        stats_frame.pack(fill=tk.X, padx=15, pady=10)
        
        if not self.results_df.empty:
            similarities = self.results_df['overall_similarity']
            mean_sim = similarities.mean()
            max_sim = similarities.max()
            min_sim = similarities.min()
            count = len(similarities)
            
            summary_text = f"Results: {count} proteins | Mean: {mean_sim:.3f} | Range: {min_sim:.3f} - {max_sim:.3f}"
        else:
            summary_text = "No similarity results available"
        
        tk.Label(stats_frame, text=summary_text,
                font=("Segoe UI", 10),
                fg=Theme.TEXT, bg=Theme.SECONDARY).pack(anchor=tk.W)
    
    def _create_buttons(self):
        # Fixed button area at bottom
        button_area = tk.Frame(self.window, bg=Theme.BG)
        button_area.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=20)
        
        button_frame = tk.Frame(button_area, bg=Theme.BG)
        button_frame.pack(anchor=tk.CENTER)
        
        ModernButton(button_frame, "💾 Export CSV", self._export_csv, "secondary").pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(button_frame, "✅ Close", self._close, "success").pack(side=tk.LEFT)
    
    def _export_csv(self):
        try:
            if self.results_df.empty:
                messagebox.showwarning("No Data", "No results to export")
                return
            
            filename = filedialog.asksaveasfilename(
                title="Save Similarity Results",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                export_df = self.results_df.copy()
                export_df.insert(0, 'Rank', range(1, len(export_df) + 1))
                export_df.to_csv(filename, index=False)
                messagebox.showinfo("Success", f"Results exported to:\n{filename}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export results:\n{e}")
    
    def _close(self):
        self.window.destroy()


class SimilarityConfigDialog:
    """Enhanced configuration dialog with preset and custom options"""
    
    def __init__(self, parent, protein_data):
        self.parent = parent
        self.protein_data = protein_data
        self.result = None
        
        self.protein_ids = protein_data['UniProt_ID'].tolist() if not protein_data.empty else []
        
        # Create dialog with larger size for sliders
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Similarity Analysis Configuration")
        self.dialog.geometry("650x800")  # Increased size for sliders
        self.dialog.configure(bg=Theme.BG)
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center on parent
        self.dialog.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() // 2) - 325
        y = parent.winfo_rooty() + (parent.winfo_height() // 2) - 400
        self.dialog.geometry(f"+{x}+{y}")
        
        # Variables
        self.central_protein_var = tk.StringVar()
        self.analysis_mode_var = tk.StringVar(value="preset")  # "preset" or "custom"
        self.preset_var = tk.StringVar(value="basic")
        
        # Slider variables for custom weights
        self.weight_vars = {}
        self.weight_categories = [
            ('sequence_length', 'Sequence Length', 0.15),
            ('molecular_weight', 'Molecular Weight', 0.20),
            ('isoelectric_point', 'Isoelectric Point', 0.15),
            ('gravy_score', 'GRAVY Score', 0.15),
            ('sequence_identity', 'Sequence Identity', 0.15),
            ('functional_keywords', 'Functional Keywords', 0.10),
            ('organism_similarity', 'Organism Similarity', 0.05),
            ('extinction_coefficient', 'Extinction Coefficient', 0.05)
        ]
        
        for key, _, default_val in self.weight_categories:
            self.weight_vars[key] = tk.DoubleVar(value=default_val)
        
        self._create_interface()
    
    def show(self):
        self.parent.wait_window(self.dialog)
        return self.result
    
    def _create_interface(self):
        """Create enhanced configuration interface"""
        # Header
        header_frame = tk.Frame(self.dialog, bg=Theme.CYAN, height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🔬 Similarity Analysis Setup",
                font=("Segoe UI", 16, "bold"),
                fg="white", bg=Theme.CYAN).pack(expand=True, pady=15)
        
        # Scrollable content area
        canvas = tk.Canvas(self.dialog, bg=Theme.BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=Theme.BG)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=(20, 0), pady=(10, 0))
        scrollbar.pack(side="right", fill="y", pady=(10, 0))
        
        # Content sections
        self._create_central_protein_section(scrollable_frame)
        self._create_analysis_mode_section(scrollable_frame)
        self._create_preset_section(scrollable_frame)
        self._create_custom_weights_section(scrollable_frame)
        self._create_data_summary(scrollable_frame)
        
        # Fixed buttons at bottom
        self._create_buttons()
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def _create_central_protein_section(self, parent):
        """Create central protein selection"""
        section_frame = tk.Frame(parent, bg=Theme.SECONDARY)
        section_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_frame = tk.Frame(section_frame, bg=Theme.TERTIARY, height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="🎯 Central Protein",
                font=("Segoe UI", 12, "bold"),
                fg=Theme.TEXT, bg=Theme.TERTIARY).pack(side=tk.LEFT, padx=15, pady=8)
        
        content_frame = tk.Frame(section_frame, bg=Theme.SECONDARY)
        content_frame.pack(fill=tk.X, padx=15, pady=15)
        
        tk.Label(content_frame, text="Select protein to compare all others against:",
                font=("Segoe UI", 10),
                fg=Theme.TEXT, bg=Theme.SECONDARY).pack(anchor=tk.W, pady=(0, 8))
        
        if self.protein_ids:
            self.central_combo = ttk.Combobox(content_frame,
                                            textvariable=self.central_protein_var,
                                            values=self.protein_ids,
                                            state="readonly",
                                            width=40)
            self.central_combo.pack(fill=tk.X)
            self.central_combo.set(self.protein_ids[0])
        else:
            tk.Label(content_frame, text="No proteins available",
                    font=("Segoe UI", 10),
                    fg=Theme.RED, bg=Theme.SECONDARY).pack()
    
    def _create_analysis_mode_section(self, parent):
        """Create analysis mode selection (preset vs custom)"""
        section_frame = tk.Frame(parent, bg=Theme.SECONDARY)
        section_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_frame = tk.Frame(section_frame, bg=Theme.TERTIARY, height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="⚙️ Analysis Mode",
                font=("Segoe UI", 12, "bold"),
                fg=Theme.TEXT, bg=Theme.TERTIARY).pack(side=tk.LEFT, padx=15, pady=8)
        
        content_frame = tk.Frame(section_frame, bg=Theme.SECONDARY)
        content_frame.pack(fill=tk.X, padx=15, pady=15)
        
        # Mode selection
        modes_frame = tk.Frame(content_frame, bg=Theme.SECONDARY)
        modes_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Radiobutton(modes_frame, text="Use Preset Configuration", 
                      variable=self.analysis_mode_var, value="preset",
                      bg=Theme.SECONDARY, fg=Theme.TEXT, selectcolor=Theme.TERTIARY,
                      activebackground=Theme.SECONDARY, activeforeground=Theme.TEXT,
                      font=("Segoe UI", 10, "bold"),
                      command=self._on_mode_change).pack(side=tk.LEFT, anchor=tk.W)
        
        tk.Radiobutton(modes_frame, text="Custom Weights", 
                      variable=self.analysis_mode_var, value="custom",
                      bg=Theme.SECONDARY, fg=Theme.TEXT, selectcolor=Theme.TERTIARY,
                      activebackground=Theme.SECONDARY, activeforeground=Theme.TEXT,
                      font=("Segoe UI", 10, "bold"),
                      command=self._on_mode_change).pack(side=tk.LEFT, anchor=tk.W, padx=(20, 0))
    
    def _create_preset_section(self, parent):
        """Create preset selection"""
        self.preset_frame = tk.Frame(parent, bg=Theme.SECONDARY)
        self.preset_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_frame = tk.Frame(self.preset_frame, bg=Theme.PURPLE, height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="📋 Analysis Presets",
                font=("Segoe UI", 12, "bold"),
                fg="white", bg=Theme.PURPLE).pack(side=tk.LEFT, padx=15, pady=8)
        
        content_frame = tk.Frame(self.preset_frame, bg=Theme.SECONDARY)
        content_frame.pack(fill=tk.X, padx=15, pady=15)
        
        presets = [
            ("basic", "Basic Analysis", "Balanced focus on core properties"),
            ("sequence", "Sequence Focus", "Emphasizes sequence-related features"),
            ("biochemical", "Biochemical Focus", "Focuses on physicochemical properties"),
            ("functional", "Functional Focus", "Emphasizes annotations and context")
        ]
        
        for value, name, description in presets:
            frame = tk.Frame(content_frame, bg=Theme.SECONDARY)
            frame.pack(fill=tk.X, pady=3)
            
            tk.Radiobutton(frame, text=name, variable=self.preset_var, value=value,
                          bg=Theme.SECONDARY, fg=Theme.TEXT, selectcolor=Theme.TERTIARY,
                          activebackground=Theme.SECONDARY, activeforeground=Theme.TEXT,
                          font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, anchor=tk.W)
            
            tk.Label(frame, text=f" - {description}",
                    font=("Segoe UI", 9),
                    fg=Theme.TEXT_MUTED, bg=Theme.SECONDARY).pack(side=tk.LEFT, anchor=tk.W)
    
    def _create_custom_weights_section(self, parent):
        """Create custom weights section with sliders"""
        self.custom_frame = tk.Frame(parent, bg=Theme.SECONDARY)
        self.custom_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_frame = tk.Frame(self.custom_frame, bg=Theme.GREEN, height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="🎛️ Custom Weights",
                font=("Segoe UI", 12, "bold"),
                fg="white", bg=Theme.GREEN).pack(side=tk.LEFT, padx=15, pady=8)
        
        # Weight total label
        self.total_label = tk.Label(title_frame, text="Total: 1.00",
                                   font=("Segoe UI", 10, "bold"),
                                   fg="white", bg=Theme.GREEN)
        self.total_label.pack(side=tk.RIGHT, padx=15, pady=8)
        
        content_frame = tk.Frame(self.custom_frame, bg=Theme.SECONDARY)
        content_frame.pack(fill=tk.X, padx=15, pady=15)
        
        # Info text
        tk.Label(content_frame, text="Adjust weights for each similarity category (total should equal 1.0):",
                font=("Segoe UI", 10),
                fg=Theme.TEXT, bg=Theme.SECONDARY).pack(anchor=tk.W, pady=(0, 10))
        
        # Create sliders for each category
        for key, label, default_val in self.weight_categories:
            self._create_weight_slider(content_frame, key, label, default_val)
        
        # Reset and normalize buttons
        button_frame = tk.Frame(content_frame, bg=Theme.SECONDARY)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ModernButton(button_frame, "Reset to Defaults", self._reset_weights, "secondary").pack(side=tk.LEFT)
        ModernButton(button_frame, "Auto-Normalize", self._normalize_weights, "primary").pack(side=tk.LEFT, padx=(10, 0))
        
        # Initially hide custom section
        self.custom_frame.pack_forget()
    
    def _create_weight_slider(self, parent, key, label, default_val):
        """Create a weight slider for a category"""
        frame = tk.Frame(parent, bg=Theme.SECONDARY)
        frame.pack(fill=tk.X, pady=3)
        
        # Label
        label_frame = tk.Frame(frame, bg=Theme.SECONDARY)
        label_frame.pack(fill=tk.X)
        
        tk.Label(label_frame, text=label, width=20, anchor="w",
                font=("Segoe UI", 9, "bold"),
                fg=Theme.TEXT, bg=Theme.SECONDARY).pack(side=tk.LEFT)
        
        # Value label
        value_label = tk.Label(label_frame, text=f"{default_val:.2f}",
                              font=("Segoe UI", 9),
                              fg=Theme.CYAN, bg=Theme.SECONDARY)
        value_label.pack(side=tk.RIGHT)
        
        # Slider
        slider = tk.Scale(frame, from_=0.0, to=1.0, resolution=0.01,
                         orient=tk.HORIZONTAL, variable=self.weight_vars[key],
                         bg=Theme.TERTIARY, fg=Theme.TEXT, highlightthickness=0,
                         troughcolor=Theme.BG, activebackground=Theme.CYAN,
                         command=lambda val, lbl=value_label: self._update_weight_label(val, lbl))
        slider.pack(fill=tk.X, pady=(2, 0))
        
        # Store reference to update label
        self.weight_vars[key].label = value_label
    
    def _update_weight_label(self, value, label):
        """Update weight label and total"""
        label.config(text=f"{float(value):.2f}")
        self._update_total()
    
    def _update_total(self):
        """Update total weight display"""
        total = sum(var.get() for var in self.weight_vars.values())
        color = Theme.GREEN if 0.98 <= total <= 1.02 else Theme.RED
        self.total_label.config(text=f"Total: {total:.2f}", fg=color)
    
    def _reset_weights(self):
        """Reset weights to default values"""
        for key, _, default_val in self.weight_categories:
            self.weight_vars[key].set(default_val)
        self._update_total()
    
    def _normalize_weights(self):
        """Normalize weights to sum to 1.0"""
        total = sum(var.get() for var in self.weight_vars.values())
        if total > 0:
            for var in self.weight_vars.values():
                var.set(var.get() / total)
        self._update_total()
    
    def _on_mode_change(self):
        """Handle analysis mode change"""
        mode = self.analysis_mode_var.get()
        if mode == "preset":
            self.custom_frame.pack_forget()
            self.preset_frame.pack(fill=tk.X, pady=(0, 15), before=self.custom_frame)
        else:  # custom
            self.preset_frame.pack_forget()
            self.custom_frame.pack(fill=tk.X, pady=(0, 15), before=self.data_summary_frame)
            self._update_total()
    
    def _create_data_summary(self, parent):
        """Create data availability summary"""
        self.data_summary_frame = tk.Frame(parent, bg=Theme.SECONDARY)
        self.data_summary_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_frame = tk.Frame(self.data_summary_frame, bg=Theme.TERTIARY, height=35)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="📊 Data Availability",
                font=("Segoe UI", 12, "bold"),
                fg=Theme.TEXT, bg=Theme.TERTIARY).pack(side=tk.LEFT, padx=15, pady=8)
        
        content_frame = tk.Frame(self.data_summary_frame, bg=Theme.SECONDARY)
        content_frame.pack(fill=tk.X, padx=15, pady=15)
        
        if not self.protein_data.empty:
            total = len(self.protein_data)
            
            fields_to_check = [
                ('sequence', 'Sequences'),
                ('mw', 'Molecular Weights'),
                ('pi', 'Isoelectric Points'),
                ('gravy', 'GRAVY Scores'),
                ('keywords', 'Keywords'),
                ('organism', 'Organisms')
            ]
            
            for field, label in fields_to_check:
                if field in self.protein_data.columns:
                    valid_count = sum(1 for v in self.protein_data[field] 
                                    if pd.notna(v) and str(v) not in ['', 'NO VALUE FOUND', 'nan'])
                    percentage = (valid_count / total) * 100 if total > 0 else 0
                    
                    color = Theme.GREEN if percentage > 50 else Theme.TEXT_MUTED if percentage > 20 else Theme.RED
                    
                    frame = tk.Frame(content_frame, bg=Theme.SECONDARY)
                    frame.pack(fill=tk.X, pady=2)
                    
                    tk.Label(frame, text=f"{label}:", width=18, anchor="w",
                            font=("Segoe UI", 9),
                            fg=Theme.TEXT, bg=Theme.SECONDARY).pack(side=tk.LEFT)
                    
                    tk.Label(frame, text=f"{valid_count}/{total} ({percentage:.0f}%)",
                            font=("Segoe UI", 9),
                            fg=color, bg=Theme.SECONDARY).pack(side=tk.LEFT)
        else:
            tk.Label(content_frame, text="No protein data available",
                    font=("Segoe UI", 10),
                    fg=Theme.RED, bg=Theme.SECONDARY).pack()
    
    def _create_buttons(self):
        """Create dialog buttons - FIXED VERSION"""
        # Create button area at the very bottom of the dialog
        button_area = tk.Frame(self.dialog, bg=Theme.BG)
        button_area.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=20)
        
        # Create centered button frame
        button_frame = tk.Frame(button_area, bg=Theme.BG)
        button_frame.pack(anchor=tk.CENTER)
        
        # Create buttons
        self.cancel_btn = ModernButton(button_frame, "Cancel", self._cancel, "secondary")
        self.cancel_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        self.start_btn = ModernButton(button_frame, "Start Analysis", self._start_analysis, "success")
        self.start_btn.pack(side=tk.LEFT)
        
        # Force geometry update
        self.dialog.update()
    
    def _start_analysis(self):
        """Start the similarity analysis"""
        if not self.protein_ids:
            messagebox.showerror("Error", "No proteins available for analysis")
            return
        
        central_protein = self.central_protein_var.get()
        if not central_protein:
            messagebox.showerror("Error", "Please select a central protein")
            return
        
        mode = self.analysis_mode_var.get()
        
        if mode == "preset":
            preset = self.preset_var.get()
            weights = None  # Will use preset weights
        else:  # custom
            # Validate custom weights
            total_weight = sum(var.get() for var in self.weight_vars.values())
            if not (0.98 <= total_weight <= 1.02):
                messagebox.showerror("Invalid Weights", 
                                   f"Custom weights must sum to 1.0 (currently {total_weight:.2f})\n"
                                   f"Use 'Auto-Normalize' button to fix this.")
                return
            
            # Get custom weights
            weights = {key: var.get() for key, var in self.weight_vars.items()}
            preset = "custom"
        
        self.result = {
            'central_protein': central_protein,
            'mode': mode,
            'preset': preset,
            'weights': weights,
            'action': 'start'
        }
        
        self.dialog.destroy()
    
    def _cancel(self):
        """Cancel the dialog"""
        self.result = {'action': 'cancel'}
        self.dialog.destroy()


class SimilarityAnalysisModal:
    """Main modal for similarity analysis"""
    
    def __init__(self, parent, output_file, protmerge_app):
        self.parent = parent
        self.output_file = output_file
        self.app = protmerge_app
        self.result = None
        
        # Load protein data
        self.protein_data = self._load_data()
    
    def show(self):
        """Show the similarity analysis workflow"""
        try:
            # Check if we have enough data
            if self.protein_data.empty or len(self.protein_data) < 2:
                messagebox.showerror(
                    "Insufficient Data",
                    f"Need at least 2 proteins for similarity analysis.\n"
                    f"Found {len(self.protein_data)} proteins."
                )
                return "insufficient_data"
            
            # Show configuration dialog
            config_dialog = SimilarityConfigDialog(self.parent, self.protein_data)
            config_result = config_dialog.show()
            
            if not config_result or config_result['action'] == 'cancel':
                return "cancelled"
            
            # Run analysis
            return self._run_analysis(config_result)
            
        except Exception as e:
            messagebox.showerror("Error", f"Similarity analysis failed:\n{e}")
            return "error"
    
    def _load_data(self):
        """Load protein data from Excel file"""
        try:
            # Load main results sheet
            main_df = pd.read_excel(self.output_file, sheet_name='ProtMerge_Results')
            
            # Convert Excel column names to internal format
            column_mapping = {
                'UniProt ID': 'UniProt_ID',
                'Organism': 'organism',
                'Gene Name': 'gene_name',
                'Protein Function/Notes': 'function',
                'Environment Source': 'environment',
                'Protein Sequence': 'sequence',
                '% Identity (Top Hit)': 'identity',
                'E-value (Top Hit)': 'evalue',
                'ProtParam: MW': 'mw',
                'ProtParam: pI': 'pi',
                'ProtParam: GRAVY': 'gravy',
                'Extinction Coefficient (M-1 cm-1)': 'ext',
                'Relevant Keywords': 'keywords'
            }
            
            # Try to load amino acid data
            try:
                aa_df = pd.read_excel(self.output_file, sheet_name='Amino_Acid_Composition')
                if 'UniProt ID' in aa_df.columns:
                    aa_cols = [col for col in aa_df.columns if col != 'UniProt ID']
                    main_df = main_df.merge(aa_df[['UniProt ID'] + aa_cols], 
                                          on='UniProt ID', how='left')
            except Exception:
                pass  # Amino acid data not available
            
            # Create results DataFrame with internal column names
            results = pd.DataFrame()
            
            for excel_col, internal_col in column_mapping.items():
                if excel_col in main_df.columns:
                    results[internal_col] = main_df[excel_col]
            
            return results
            
        except Exception as e:
            messagebox.showerror("Data Loading Error", f"Could not load protein data:\n{e}")
            return pd.DataFrame()
    
    def _run_analysis(self, config):
        """Run the similarity analysis"""
        central_protein = config['central_protein']
        mode = config['mode']
        
        # Get weights based on mode
        if mode == "preset":
            preset = config['preset']
            preset_methods = {
                'basic': SimilarityPresets.get_basic_preset,
                'sequence': SimilarityPresets.get_sequence_preset,
                'biochemical': SimilarityPresets.get_biochemical_preset,
                'functional': SimilarityPresets.get_functional_preset
            }
            weights = preset_methods.get(preset, SimilarityPresets.get_basic_preset)()
        else:  # custom
            weights = config['weights']
        
        # Show progress dialog
        progress_dialog = SimilarityProgressDialog(self.parent)
        
        # Run analysis in thread
        analysis_thread = threading.Thread(
            target=self._analysis_worker,
            args=(central_protein, weights, progress_dialog),
            daemon=True
        )
        analysis_thread.start()
        
        # Wait for completion
        while analysis_thread.is_alive() and not progress_dialog.cancelled:
            self.parent.update()
            time.sleep(0.1)
        
        if progress_dialog.cancelled:
            progress_dialog.close()
            return "cancelled"
        
        return "complete"
    
    def _analysis_worker(self, central_protein, weights, progress_dialog):
        """Worker thread for similarity analysis"""
        try:
            def progress_callback(percentage, text):
                """Update progress dialog"""
                try:
                    progress_dialog.update_progress(percentage, text)
                except:
                    pass  # Dialog might be closed
            
            # Run analysis
            progress_callback(10, "Initializing similarity analysis...")
            
            results_df = run_similarity_analysis(
                self.protein_data,
                central_protein,
                weights,
                progress_callback
            )
            
            progress_callback(100, "Analysis complete!")
            
            # Close progress dialog and show results
            self.parent.after(500, lambda: self._show_results(results_df, central_protein, progress_dialog))
            
        except Exception as e:
            # Show error and close progress dialog
            self.parent.after(0, lambda: self._show_error(str(e), progress_dialog))
    
    def _show_results(self, results_df, central_protein, progress_dialog):
        """Show analysis results"""
        try:
            progress_dialog.close()
            
            if results_df.empty:
                messagebox.showwarning("No Results", "Similarity analysis produced no results.")
                return
            
            # Save results to Excel
            self._save_results_to_excel(results_df, central_protein)
            
            # Show results viewer
            results_viewer = SimilarityResultsViewer(self.parent, results_df, central_protein)
            
            messagebox.showinfo(
                "Analysis Complete",
                f"Similarity analysis completed successfully!\n\n"
                f"Analyzed {len(results_df)} proteins relative to {central_protein}.\n"
                f"Results have been saved to the Excel file."
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to display results:\n{e}")
    
    def _show_error(self, error_msg, progress_dialog):
        """Show error message"""
        progress_dialog.close()
        messagebox.showerror("Analysis Failed", f"Similarity analysis failed:\n\n{error_msg}")
    
    def _save_results_to_excel(self, results_df, central_protein):
        """Save similarity results to Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load workbook
            wb = load_workbook(self.output_file)
            
            # Remove existing similarity sheet if present
            if 'Similarity_Analysis' in wb.sheetnames:
                wb.remove(wb['Similarity_Analysis'])
            
            # Create new sheet
            ws = wb.create_sheet('Similarity_Analysis')
            
            # Prepare data with metadata
            export_data = pd.DataFrame({
                'Rank': ['Analysis Info:', '', 'Central Protein:', f'{central_protein}', '', 'Results:'] + 
                        list(range(1, len(results_df) + 1)),
                'Protein ID': ['', '', 'Date:', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'), '', ''] +
                              results_df['protein_id'].tolist(),
                'Overall Similarity': ['', '', 'Total Proteins:', str(len(results_df)), '', ''] +
                                     [f"{sim:.4f}" for sim in results_df['overall_similarity']],
                'Data Quality': ['', '', 'Analysis Type:', 'Similarity Matrix', '', ''] +
                               [f"{qual:.3f}" for qual in results_df['data_quality']]
            })
            
            # Write to sheet
            for r in dataframe_to_rows(export_data, index=False, header=True):
                ws.append(r)
            
            # Save workbook
            wb.save(self.output_file)
            
        except Exception as e:
            # Non-critical error - analysis still succeeded
            messagebox.showwarning("Save Warning", f"Could not save results to Excel:\n{e}")


# =============================================================================
# INTEGRATION FUNCTION
# =============================================================================

def launch_similarity_analysis(parent, output_file, protmerge_app):
    """
    Launch similarity analysis modal - main integration function
    
    Args:
        parent: Parent GUI window
        output_file: Path to Excel file with protein data
        protmerge_app: ProtMerge application instance
    
    Returns:
        String indicating result: "complete", "cancelled", "error", etc.
    """
    try:
        modal = SimilarityAnalysisModal(parent, output_file, protmerge_app)
        return modal.show()
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to launch similarity analysis:\n{e}")
        return "error"


# =============================================================================
# TESTING FUNCTION
# =============================================================================

def test_similarity_gui():
    """Test the similarity GUI components"""
    root = tk.Tk()
    root.withdraw()  # Hide main window
    
    # Create test data
    test_data = pd.DataFrame({
        'UniProt_ID': ['P12345', 'Q67890', 'R11111', 'S22222', 'T33333'],
        'organism': ['Human', 'Mouse', 'Rat', 'Dog', 'Cat'],
        'sequence': ['ACDEF' * 10, 'BCDEF' * 12, 'CDEFG' * 8, 'DEFGH' * 15, 'EFGHI' * 9],
        'mw': [50000, 60000, 70000, 45000, 55000],
        'pi': [7.0, 8.0, 9.0, 6.5, 7.5],
        'gravy': [-0.5, 0.0, 0.5, -0.2, 0.3],
        'keywords': ['enzyme; binding', 'transport; membrane', 'catalysis', 'signaling; receptor', 'metabolic']
    })
    
    # Test configuration dialog
    config_dialog = SimilarityConfigDialog(root, test_data)
    result = config_dialog.show()
    
    print(f"Configuration result: {result}")
    
    root.destroy()


if __name__ == "__main__":
    # Test the GUI components
    test_similarity_gui()