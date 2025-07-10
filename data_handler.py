"""
Data loading and processing for ProtMerge v1.2.0
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import logging
from config import OUTPUT_COLUMNS, AMINO_ACID_COLUMNS

class DataHandler:
    """Handles data loading and processing operations"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def load_excel_data(self, input_file, sheet_name, column_index, safe_mode=True):
        """Load and prepare data from Excel file"""
        try:
            # Read Excel file
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            
            # Get UniProt column
            uniprot_col_name = df.columns[column_index]
            uniprot_ids = df[uniprot_col_name].dropna().tolist()
            
            # Create results DataFrame
            results = pd.DataFrame({
                'UniProt_ID': uniprot_ids,
                'Input_Row': [i for i, val in enumerate(df[uniprot_col_name]) if pd.notna(val)]
            })
            
            # Initialize columns
            self._initialize_columns(results, input_file, safe_mode)
            
            self.logger.info(f"Loaded {len(uniprot_ids)} UniProt IDs from '{uniprot_col_name}'")
            
            return {
                'results': results,
                'input_file': Path(input_file),
                'input_sheet': sheet_name,
                'uniprot_column': uniprot_col_name
            }
            
        except Exception as e:
            self.logger.error(f"Failed to load data: {e}")
            raise
    
    def _initialize_columns(self, results, input_file, safe_mode):
        """Initialize result columns"""
        if safe_mode:
            self._load_existing_results(results, input_file)
        else:
            for key in OUTPUT_COLUMNS:
                results[key] = "NO VALUE FOUND"
    
    def _load_existing_results(self, results, input_file):
        """Load existing results if available (safe mode)"""
        try:
            existing_wb = load_workbook(input_file)
            if 'ProtMerge_Results' in existing_wb.sheetnames:
                self.logger.info("Loading existing data...")
                
                existing_df = pd.read_excel(input_file, sheet_name='ProtMerge_Results')
                
                # Initialize columns
                for key in OUTPUT_COLUMNS:
                    results[key] = "NO VALUE FOUND"
                
                # Load existing data
                for idx, row in results.iterrows():
                    existing_row = existing_df[existing_df['UniProt ID'] == row['UniProt_ID']]
                    if not existing_row.empty:
                        self._load_existing_row(results, idx, existing_row.iloc[0])
                        
                self.logger.info("Existing data loaded")
            else:
                for key in OUTPUT_COLUMNS:
                    results[key] = "NO VALUE FOUND"
                    
        except Exception as e:
            self.logger.warning(f"Could not load existing results: {e}")
            for key in OUTPUT_COLUMNS:
                results[key] = "NO VALUE FOUND"
    
    def _load_existing_row(self, results, idx, existing_row):
        """Load data for single row from existing results"""
        for key, column_name in OUTPUT_COLUMNS.items():
            if column_name in existing_row.index:
                existing_value = existing_row[column_name]
                if self._is_valid_value(existing_value):
                    results.at[idx, key] = existing_value
    
    def _is_valid_value(self, value):
        """Check if existing value is valid"""
        return (pd.notna(value) and 
                str(value).strip() != '' and 
                str(value) != "NO VALUE FOUND")

    def should_update_field(self, results, idx, field, safe_mode=True):
        """Check if field should be updated"""
        if not safe_mode:
            return True
        current_value = results.at[idx, field]
        return (pd.isna(current_value) or 
                str(current_value).strip() == '' or 
                str(current_value) == "NO VALUE FOUND")