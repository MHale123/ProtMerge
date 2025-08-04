"""
excel_formatter.py for ProtMerge v1.2.0 - FIXED: UniProt sheet bug and improved logic
"""

import logging
import time
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import OUTPUT_COLUMNS, AMINO_ACID_COLUMNS, PDB_COLUMNS, SIMILARITY_COLUMNS, HUMAN_PROTEIN_COLUMNS, THEMES

class ExcelFormatter:
    """
    Handles Excel output formatting and saving with professional styling.
    FIXED: Only creates sheets when analysis was requested AND actual data is available.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.logger.info("ExcelFormatter initialized")
    
    def save_results(self, input_file, results, options):
        """
        Save results to professionally formatted Excel file with multiple sheets.
        FIXED: Only creates main sheet if UniProt-dependent analyses were requested AND have data.
        """
        try:
            self.logger.info(f"Creating Excel output for {len(results)} proteins")
        
            # Prepare output file path
            output_file = self._prepare_output_file(input_file)
        
            # Create workbook
            wb = self._create_workbook(input_file)
        
            # Check if we need a main sheet (only if UniProt-dependent analyses were requested)
            needs_main_sheet = self._needs_main_sheet(options)
            sheets_created = []
        
            # Only create main sheet if UniProt-dependent analyses were requested AND have data
            if needs_main_sheet and self._has_main_sheet_data(results, options):
                self._create_main_sheet(wb, results, options)
                sheets_created.append('ProtMerge_Results')
            elif needs_main_sheet:
                self.logger.info("UniProt-dependent analyses requested but no data available - skipping main sheet")
        
            # Create optional sheets only if they have actual data AND were requested
            if options.get('amino_acid', False) and self._has_amino_acid_data(results):
                self._create_amino_acid_sheet(wb, results)
                sheets_created.append('Amino_Acid_Composition')
        
            if options.get('pdb_search', False) and self._has_pdb_data(results):
                self._create_pdb_sheet(wb, results)
                sheets_created.append('PDB_Structural_Analysis')
        
            # Human analyses sheet - create if human analyses were requested
            human_analyses_requested = options.get('compartments', False) or options.get('hpa', False)
            if human_analyses_requested:
                if self._has_human_protein_data(results):
                    self._create_human_protein_sheet(wb, results)
                    sheets_created.append('Human_Protein_Data')
                else:
                    # Create minimal sheet even if no data, since this was the primary analysis requested
                    self._create_minimal_human_sheet(wb, results, options)
                    sheets_created.append('Human_Protein_Data')

            # Similarity sheet if results are available
            if options.get('similarity_results') is not None:
                self._create_similarity_sheet(wb, results, options)
                sheets_created.append('Similarity_Analysis')
        
            # If no sheets were created, create a minimal results sheet
            if not sheets_created:
                self._create_minimal_results_sheet(wb, results)
                sheets_created.append('Analysis_Results')
        
            # Save workbook
            saved_file = self._save_workbook(wb, output_file)
            if saved_file:
                self.logger.info(f"Created Excel sheets: {', '.join(sheets_created)}")
        
            return saved_file
        
        except Exception as e:
            self.logger.error(f"Excel formatting failed: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return self._create_emergency_backup(results, input_file)
    
    def _needs_main_sheet(self, options):
        """Check if main sheet with UniProt-dependent data is needed"""
        uniprot_dependent_analyses = ['uniprot', 'protparam', 'blast', 'pdb_search']
        return any(options.get(analysis, False) for analysis in uniprot_dependent_analyses)

    def _has_main_sheet_data(self, results, options):
        """Check if we have data for the main sheet"""
        if not self._needs_main_sheet(options):
            return False
    
        # Check if any requested UniProt-dependent analyses have data
        if options.get('uniprot', False) and self._has_uniprot_data(results):
            return True
        if options.get('protparam', False) and self._has_protparam_data(results):
            return True
        if options.get('blast', False) and self._has_blast_data(results):
            return True
        if options.get('pdb_search', False) and self._has_pdb_data(results):
            return True
    
        return False

    def _create_minimal_human_sheet(self, wb, results, options):
        """Create minimal sheet for human-only analysis"""
        if 'Human_Protein_Data' in wb.sheetnames:
            wb.remove(wb['Human_Protein_Data'])
    
        ws = wb.create_sheet('Human_Protein_Data')
    
        # Create minimal dataframe with gene IDs (which are in UniProt_ID column at this point)
        human_df = pd.DataFrame()
        human_df['Gene ID'] = results['UniProt_ID']  # These are actually gene IDs for human-only analysis
    
        # Add human protein columns
        for internal_key, excel_column in HUMAN_PROTEIN_COLUMNS.items():
            if internal_key in results.columns:
                human_df[excel_column] = results[internal_key]
            else:
                human_df[excel_column] = "NO VALUE FOUND"
    
        self._write_dataframe_to_sheet(ws, human_df)
        self._format_sheet(ws, THEMES['human'])
    
        self.logger.info("Created minimal human protein data sheet")

    def _create_minimal_results_sheet(self, wb, results):
        """Create minimal results sheet when no other sheets are appropriate"""
        if 'Analysis_Results' in wb.sheetnames:
            wb.remove(wb['Analysis_Results'])
    
        ws = wb.create_sheet('Analysis_Results')
    
        # Create basic dataframe
        minimal_df = pd.DataFrame()
        minimal_df['Input ID'] = results['UniProt_ID']
    
        # Add original gene ID if available
        if 'Original_Gene_ID' in results.columns:
            has_gene_data = results['Original_Gene_ID'].apply(
                lambda x: pd.notna(x) and str(x).strip() != ''
            ).any()
        
            if has_gene_data:
                minimal_df['Original Gene ID'] = results['Original_Gene_ID']
    
        self._write_dataframe_to_sheet(ws, minimal_df)
        self._format_sheet(ws, THEMES['main'])
    
        self.logger.info("Created minimal analysis results sheet")

    def _determine_columns_to_include(self, results, options):
        """
        FIXED: Determine which columns to include based on what was actually analyzed AND requested
        This method now properly checks both conditions to avoid the UniProt sheet bug
        """
        columns_to_include = {}
        
        # FIXED: UniProt data columns (only if UniProt analysis was requested AND has data)
        if options.get('uniprot', False):
            if self._has_uniprot_data(results):
                uniprot_columns = {
                    'organism': 'Organism',
                    'gene_name': 'Gene Name', 
                    'function': 'Protein Function/Notes',
                    'environment': 'Environment Source',
                    'sequence': 'Protein Sequence',
                    'alphafold': 'AlphaFold Link',
                    'keywords': 'Relevant Keywords',
                    'structure': 'Structure Type'
                }
                columns_to_include.update(uniprot_columns)
            else:
                self.logger.info("UniProt was requested but no data was found - skipping UniProt columns")
        
        # FIXED: ProtParam columns (only if ProtParam analysis was requested AND has data)
        if options.get('protparam', False):
            if self._has_protparam_data(results):
                protparam_columns = {
                    'mw': 'ProtParam: MW',
                    'pi': 'ProtParam: pI',
                    'gravy': 'ProtParam: GRAVY',
                    'ext': 'Extinction Coefficient (M-1 cm-1)'
                }
                columns_to_include.update(protparam_columns)
            else:
                self.logger.info("ProtParam was requested but no data was found - skipping ProtParam columns")
        
        # FIXED: BLAST columns (only if BLAST analysis was requested AND has data)
        if options.get('blast', False):
            if self._has_blast_data(results):
                blast_columns = {
                    'similar': 'BLAST Similar Proteins',
                    'identity': '% Identity (Top Hit)',
                    'evalue': 'E-value (Top Hit)',
                    'align': 'Alignment Length (Top Hit)'
                }
                columns_to_include.update(blast_columns)
            else:
                self.logger.info("BLAST was requested but no data was found - skipping BLAST columns")
        
        return columns_to_include
    
    def _prepare_output_file(self, input_file):
        """Prepare output file path with conflict resolution"""
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}_ProtMerge.xlsx"
        
        # Handle file conflicts
        counter = 1
        while output_file.exists():
            try:
                # Test if file is accessible
                with open(output_file, 'r+b'):
                    break
            except PermissionError:
                # File is open, try alternative name
                output_file = input_path.parent / f"{input_path.stem}_ProtMerge_{counter}.xlsx"
                counter += 1
                if counter > 10:
                    break
        
        return output_file
    
    def _create_workbook(self, input_file):
        """Create or load workbook"""
        try:
            wb = load_workbook(input_file)
            self.logger.debug("Loaded existing workbook")
        except Exception:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            self.logger.debug("Created new workbook")
        
        return wb
    
    def _create_main_sheet(self, wb, results, options):
        """Create main results sheet with only data that was actually collected"""
        # Remove existing sheet if it exists
        if 'ProtMerge_Results' in wb.sheetnames:
            wb.remove(wb['ProtMerge_Results'])
        
        ws = wb.create_sheet('ProtMerge_Results')
        
        # Prepare data for output - only include columns that were actually analyzed
        output_df = self._prepare_main_data_selective(results, options)
        
        # Write to sheet
        self._write_dataframe_to_sheet(ws, output_df)
        self._format_sheet(ws, THEMES['main'])
        
        self.logger.info(f"Created main sheet with {len(output_df)} entries and {len(output_df.columns)} data columns")
    
    def _prepare_main_data_selective(self, results, options):
        """FIXED: Prepare main results data for Excel output - only include analyzed data"""
        output_df = pd.DataFrame()
        
        # Always include UniProt ID
        output_df['UniProt ID'] = results['UniProt_ID']
    
        # Add Original Gene ID column if it exists and has data
        if 'Original_Gene_ID' in results.columns:
            has_gene_data = results['Original_Gene_ID'].apply(
                lambda x: pd.notna(x) and str(x).strip() != '' and str(x) != str(results['UniProt_ID'].iloc[0])
            ).any()
        
            if has_gene_data:
                output_df['Original Gene ID'] = results['Original_Gene_ID']
    
        # FIXED: Add columns based on what was actually analyzed AND requested
        columns_to_include = self._determine_columns_to_include(results, options)
        
        for internal_key, excel_column in columns_to_include.items():
            if internal_key in results.columns:
                output_df[excel_column] = results[internal_key]
            else:
                # This shouldn't happen with the fixed logic, but keeping as safety net
                self.logger.warning(f"Column {internal_key} was expected but not found in results")
                output_df[excel_column] = "NO VALUE FOUND"
    
        self.logger.info(f"Main sheet will include columns: {list(columns_to_include.values())}")
        return output_df
    
    def _has_uniprot_data(self, results):
        """Check if UniProt data is actually present"""
        uniprot_keys = ['organism', 'gene_name', 'function', 'sequence']
        return self._has_data_for_fields(results, uniprot_keys)
    
    def _has_protparam_data(self, results):
        """Check if ProtParam data is actually present"""
        protparam_keys = ['mw', 'pi', 'gravy', 'ext']
        return self._has_data_for_fields(results, protparam_keys)
    
    def _has_blast_data(self, results):
        """Check if BLAST data is actually present"""
        blast_keys = ['similar', 'identity', 'evalue', 'align']
        return self._has_data_for_fields(results, blast_keys)
    
    def _has_amino_acid_data(self, results):
        """Check if amino acid data is available"""
        aa_keys = list(AMINO_ACID_COLUMNS.keys())
        return self._has_data_for_fields(results, aa_keys)
    
    def _has_pdb_data(self, results):
        """Check if PDB data is available"""
        pdb_keys = list(PDB_COLUMNS.keys())
        return self._has_data_for_fields(results, pdb_keys)
    
    def _has_human_protein_data(self, results):
        """Check if human protein data is available"""
        human_keys = list(HUMAN_PROTEIN_COLUMNS.keys())
        return self._has_data_for_fields(results, human_keys)
    
    def _has_data_for_fields(self, results, field_keys):
        """Check if any of the specified fields have actual data"""
        for key in field_keys:
            if key in results.columns:
                has_data = results[key].apply(
                    lambda x: (pd.notna(x) and 
                              str(x) != "NO VALUE FOUND" and 
                              str(x).strip() != "" and
                              str(x).lower() != "nan")
                ).any()
                if has_data:
                    self.logger.debug(f"Found data for field: {key}")
                    return True
        self.logger.debug(f"No data found for any fields in: {field_keys}")
        return False
    
    def _create_amino_acid_sheet(self, wb, results):
        """Create amino acid composition sheet"""
        if 'Amino_Acid_Composition' in wb.sheetnames:
            wb.remove(wb['Amino_Acid_Composition'])
        
        ws = wb.create_sheet('Amino_Acid_Composition')
        
        # Prepare amino acid data
        aa_df = self._prepare_amino_acid_data(results)
        
        # Write to sheet
        self._write_dataframe_to_sheet(ws, aa_df)
        self._format_sheet(ws, THEMES['amino'])
        
        self.logger.info("Created amino acid composition sheet with actual data")

    def _create_human_protein_sheet(self, wb, results):
        """Create human protein analysis sheet"""
        if 'Human_Protein_Data' in wb.sheetnames:
            wb.remove(wb['Human_Protein_Data'])
    
        ws = wb.create_sheet('Human_Protein_Data')
    
        # Prepare human protein data
        human_df = self._prepare_human_protein_data(results)
    
        # Write to sheet
        self._write_dataframe_to_sheet(ws, human_df)
        self._format_sheet(ws, THEMES['human'])
    
        self.logger.info("Created human protein analysis sheet with actual data")
    
    def _prepare_amino_acid_data(self, results):
        """Prepare amino acid composition data"""
        aa_df = pd.DataFrame()
        aa_df['UniProt ID'] = results['UniProt_ID']
        
        # Add gene name if available
        if 'gene_name' in results.columns:
            aa_df['Gene Name'] = results['gene_name']
        else:
            aa_df['Gene Name'] = "N/A"
        
        # Add amino acid columns
        for internal_key, excel_column in AMINO_ACID_COLUMNS.items():
            if internal_key in results.columns:
                aa_df[excel_column] = results[internal_key]
            else:
                aa_df[excel_column] = "NO VALUE FOUND"
        
        return aa_df
    
    def _prepare_human_protein_data(self, results):
        """Prepare human protein data for Excel output"""
        human_df = pd.DataFrame()
        human_df['UniProt ID'] = results['UniProt_ID']
    
        # Add original gene ID if available
        if 'Original_Gene_ID' in results.columns:
            has_gene_data = results['Original_Gene_ID'].apply(
                lambda x: pd.notna(x) and str(x).strip() != ''
            ).any()
        
            if has_gene_data:
                human_df['Original Gene ID'] = results['Original_Gene_ID']
    
        # Add gene name if available
        if 'gene_name' in results.columns:
            human_df['Gene Name'] = results['gene_name']
        else:
            human_df['Gene Name'] = "N/A"
    
        # Add human protein columns
        for internal_key, excel_column in HUMAN_PROTEIN_COLUMNS.items():
            if internal_key in results.columns:
                human_df[excel_column] = results[internal_key]
            else:
                # Set appropriate default values for missing data
                if 'confidence' in internal_key and 'primary' in internal_key:
                    human_df[excel_column] = 0  # Numerical 0 for missing confidence
                elif 'reliability' in internal_key:
                    human_df[excel_column] = 0  # Numerical 0 for missing reliability
                else:
                    human_df[excel_column] = "NO VALUE FOUND"
    
        return human_df
    
    def _create_pdb_sheet(self, wb, results):
        """Create PDB structural analysis sheet"""
        if 'PDB_Structural_Analysis' in wb.sheetnames:
            wb.remove(wb['PDB_Structural_Analysis'])
        
        ws = wb.create_sheet('PDB_Structural_Analysis')
        
        # Prepare PDB data
        pdb_df = self._prepare_pdb_data(results)
        
        # Write to sheet
        self._write_dataframe_to_sheet(ws, pdb_df)
        self._format_sheet(ws, THEMES['pdb'])
        
        self.logger.info("Created PDB structural analysis sheet with actual data")
    
    def _prepare_pdb_data(self, results):
        """Prepare PDB structural data"""
        pdb_df = pd.DataFrame()
        pdb_df['UniProt ID'] = results['UniProt_ID']
        
        # Add gene name if available
        if 'gene_name' in results.columns:
            pdb_df['Gene Name'] = results['gene_name']
        else:
            pdb_df['Gene Name'] = "N/A"
        
        # Add PDB columns
        for internal_key, excel_column in PDB_COLUMNS.items():
            if internal_key in results.columns:
                pdb_df[excel_column] = results[internal_key]
            else:
                pdb_df[excel_column] = "NO VALUE FOUND"
        
        return pdb_df
    
    def _create_similarity_sheet(self, wb, results, options):
        """Create similarity analysis sheet"""
        if 'Similarity_Analysis' in wb.sheetnames:
            wb.remove(wb['Similarity_Analysis'])
        
        ws = wb.create_sheet('Similarity_Analysis')
        
        # Get similarity results from options
        similarity_results = options.get('similarity_results', None)
        central_protein = options.get('central_protein_id', 'Unknown')
        
        if similarity_results is not None and not similarity_results.empty:
            # Create comprehensive similarity data
            sim_df = self._prepare_similarity_data(similarity_results, central_protein)
        else:
            # Create placeholder data for failed analysis
            sim_df = self._create_similarity_placeholder(central_protein)
        
        # Write to sheet
        self._write_dataframe_to_sheet(ws, sim_df)
        self._format_sheet(ws, THEMES['similarity'])
        
        # Add conditional formatting for similarity scores
        self._add_similarity_formatting(ws)
        
        has_data = similarity_results is not None and not similarity_results.empty
        self.logger.info(f"Created similarity sheet ({'with data' if has_data else 'no data available'})")
    
    def _prepare_similarity_data(self, similarity_results, central_protein):
        """Prepare similarity results data for Excel"""
        # Create main results data
        sim_df = pd.DataFrame()
        
        # Add ranking and core information
        sim_df['Rank'] = range(1, len(similarity_results) + 1)
        sim_df['Protein ID'] = similarity_results['protein_id']
        sim_df['Overall Similarity'] = similarity_results['overall_similarity'].round(4)
        sim_df['Data Quality'] = similarity_results['data_quality'].round(3)
        
        # Add category breakdowns if available
        category_columns = [
            'sequence_length', 'sequence_identity', 'molecular_weight',
            'isoelectric_point', 'gravy_score', 'functional_keywords',
            'organism_similarity', 'amino_acid_composition'
        ]
        
        for col in category_columns:
            if col in similarity_results.columns:
                display_name = col.replace('_', ' ').title()
                sim_df[display_name] = similarity_results[col].round(3)
        
        # Add metadata header rows
        metadata_rows = pd.DataFrame({
            'Rank': ['Analysis Info:', 'Central Protein:', 'Analysis Date:', 'Total Proteins:', ''],
            'Protein ID': ['Similarity Analysis Results', central_protein, 
                          pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"), 
                          str(len(similarity_results)), 'Results:'],
            'Overall Similarity': ['', '', '', '', ''],
            'Data Quality': ['', '', '', '', '']
        })
        
        # Combine metadata and results
        final_df = pd.concat([metadata_rows, sim_df], ignore_index=True)
        
        return final_df
    
    def _create_similarity_placeholder(self, central_protein):
        """Create placeholder data when similarity analysis fails"""
        return pd.DataFrame({
            'Rank': ['Analysis Failed'],
            'Protein ID': [f'Central Protein: {central_protein}'],
            'Overall Similarity': ['No results available'],
            'Data Quality': ['Check analysis parameters']
        })
    
    def _add_similarity_formatting(self, ws):
        """Add conditional formatting for similarity scores"""
        try:
            from openpyxl.formatting.rule import ColorScaleRule
            
            # Find the Overall Similarity column
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Overall Similarity':
                    col_letter = chr(64 + col_idx)  # Convert to letter
                    
                    # Apply color scale from red (low) to green (high)
                    color_scale = ColorScaleRule(
                        start_type='min', start_color='FFFF0000',  # Red
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF00',  # Yellow
                        end_type='max', end_color='FF00FF00'  # Green
                    )
                    
                    ws.conditional_formatting.add(f'{col_letter}6:{col_letter}{ws.max_row}', color_scale)
                    break
                    
        except ImportError:
            # openpyxl conditional formatting not available
            self.logger.warning("Conditional formatting not available")
        except Exception as e:
            self.logger.warning(f"Could not add similarity formatting: {e}")
    
    def _write_dataframe_to_sheet(self, ws, df):
        """Write DataFrame to worksheet"""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
    def _format_sheet(self, ws, theme):
        """Apply professional formatting to worksheet"""
        if not ws.max_row or ws.max_row < 1:
            return
        
        # Define styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=theme['header'], end_color=theme['header'], fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(name="Segoe UI", size=10)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alt_fill = PatternFill(start_color=theme['alt_row'], end_color=theme['alt_row'], fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E1E5E9'),
            right=Side(style='thin', color='E1E5E9'),
            top=Side(style='thin', color='E1E5E9'),
            bottom=Side(style='thin', color='E1E5E9')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Format data rows
        for row_num in range(2, ws.max_row + 1):
            fill = alt_fill if row_num % 2 == 0 else PatternFill()
            
            for cell in ws[row_num]:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                cell.fill = fill
        
        # Auto-size columns
        self._auto_size_columns(ws)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Make hyperlinks clickable
        self._make_hyperlinks_clickable(ws)
    
    def _auto_size_columns(self, ws):
        """Auto-size columns with reasonable limits"""
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                if cell.value:
                    try:
                        length = len(str(cell.value))
                        max_length = max(max_length, length)
                    except:
                        pass
            
            # Set width with limits
            width = min(max(max_length + 2, 12), 45)
            ws.column_dimensions[column_letter].width = width
    
    def _make_hyperlinks_clickable(self, ws):
        """Make URLs clickable"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('https://'):
                    cell.hyperlink = cell.value
                    cell.font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
    
    def _save_workbook(self, wb, output_file):
        """Save workbook with error handling"""
        try:
            wb.save(output_file)
            return output_file
            
        except PermissionError:
            # Try with timestamp
            timestamp = int(time.time())
            timestamped_file = output_file.parent / f"{output_file.stem}_{timestamp}.xlsx"
            try:
                wb.save(timestamped_file)
                self.logger.info(f"Results saved to timestamped file: {timestamped_file}")
                return timestamped_file
            except Exception as e:
                self.logger.error(f"Timestamped save failed: {e}")
                return None
        
        except Exception as e:
            self.logger.error(f"Save failed: {e}")
            return None
    
    def _create_emergency_backup(self, results, input_file):
        """Create emergency backup when main save fails"""
        try:
            # Find writable directory
            backup_dirs = [
                Path.home() / "Desktop",
                Path.home() / "Documents", 
                Path.cwd()
            ]
            
            backup_dir = None
            for directory in backup_dirs:
                if directory.exists():
                    try:
                        # Test write access
                        test_file = directory / "test_write.tmp"
                        test_file.touch()
                        test_file.unlink()
                        backup_dir = directory
                        break
                    except:
                        continue
            
            if backup_dir is None:
                backup_dir = Path.cwd()
            
            backup_file = backup_dir / f"ProtMerge_Emergency_Backup_{int(time.time())}.csv"
            
            # Create simple CSV backup
            backup_df = pd.DataFrame()
            backup_df['UniProt_ID'] = results['UniProt_ID']
            
            # Add all available columns
            for col in results.columns:
                if col != 'UniProt_ID':
                    backup_df[col] = results[col]
            
            backup_df.to_csv(backup_file, index=False)
            self.logger.info(f"Emergency backup saved to: {backup_file}")
            return backup_file
            
        except Exception as e:
            self.logger.error(f"Emergency backup failed: {e}")
            return None
    
    def add_similarity_results_to_options(self, options, similarity_results, central_protein_id):
        """Add similarity results to options for Excel export."""
        options = options.copy()
        options['similarity_results'] = similarity_results
        options['central_protein_id'] = central_protein_id
        options['similarity_analysis'] = True
        
        return options